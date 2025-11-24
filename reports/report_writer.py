import json
import csv
import os
from datetime import datetime
from openpyxl import Workbook
from fpdf import FPDF
from pathlib import Path

def generate_report_files(summary, adapter_name, entity, migration_type):
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"migration_api_{adapter_name}_{timestamp}"
    output_dir = Path(__file__).resolve().parent.parent / "auditreports"
    output_dir.mkdir(parents=True, exist_ok=True)

    csv_path = output_dir / f"{base_name}.csv"

    write_csv(summary["rows"], csv_path)

    return {
        "csv": csv_path.name
    }

def write_csv(rows, path):
    if not rows:
        return
    fieldnames = sorted({key for row in rows if isinstance(row, dict) for key in row.keys()})
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})

def write_xlsx(rows, path):
    if not rows:
        print(f"⚠️ [write_xlsx] No rows to write to: {path}")
        return

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Migration Results"

    headers = list(rows[0].keys())
    ws.append(headers)

    for i, row in enumerate(rows, start=1):
        safe_row = []
        for h in headers:
            value = row.get(h, "")
            if isinstance(value, (dict, list)):
                value = json.dumps(value)  # flatten nested structures
            elif value is None:
                value = ""
            safe_row.append(str(value))  # ensure string-safe
        try:
            ws.append(safe_row)
        except Exception as e:
            print(f"❌ [write_xlsx] Failed to write row {i}: {e}")
            print(f"🔍 Row content: {safe_row}")

    try:
        wb.save(path)
        print(f"✅ [write_xlsx] Excel file saved: {path}")
    except Exception as e:
        print(f"❌ [write_xlsx] Failed to save Excel file: {e}")

def write_pdf(rows, path):
    if not rows:
        return
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=10)

    headers = sorted({key for row in rows if isinstance(row, dict) for key in row.keys()})
    col_width = 190 / len(headers)

    # Header
    for header in headers:
        pdf.cell(col_width, 10, header, border=1)
    pdf.ln()

    # Rows
    for row in rows:
        for header in headers:
            cell = str(row.get(header, ""))
            pdf.cell(col_width, 10, cell[:120], border=1)  # allow more characters
        pdf.ln()

    pdf.output(path)